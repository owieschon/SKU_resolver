"""Catalog ingestion pathways — get an unknown catalog into the decoder.

The grammar decoder (`grammar_induction.decode_catalog`) wants a uniform shape:
a list of `{sku, description}` rows. Real catalogs arrive as PDFs, Excel
workbooks, ERP item tables, or web pages. This module is the thin ingestion
layer that turns each of those into that uniform shape — so one decoder serves
every source.

Design: the row-extraction *logic* for each format is a pure function over
already-loaded content (text lines / html / a worksheet), independently
testable with no file or network. The `*CatalogSource` classes are the I/O
adapters that load the content and call those pure functions. Adding a new
source (a vendor portal scrape, a CSV drop) is a new `rows_from_*` function
plus a tiny adapter — the decoder never changes.

Format support:
  - PDF   : `rows_from_catalog_lines` over extracted text (pypdf, `[pdf]` extra,
            or any text extractor — layout-mode `pdftotext` gives the best
            columnar fidelity). Validated on a real 1,200-SKU vendor catalog.
  - Excel : `rows_from_worksheet` over openpyxl rows (openpyxl is a core dep).
  - Web   : `rows_from_html_tables` over an HTML string (stdlib html.parser).
"""
from __future__ import annotations

import re
from html.parser import HTMLParser
from typing import Protocol

# A part-number-like token: starts alphanumeric, then alphanumerics/-/./ ,
# at least one digit somewhere (pure words are not SKUs).
_CODE = re.compile(r'^[A-Za-z0-9][A-Za-z0-9\-/.]{3,}$')
# Header words that name the SKU column / description column, by source.
_SKU_HEADERS = ('part', 'sku', 'item', 'number', 'model', 'no', 'cat', '#')
_DESC_HEADERS = ('desc', 'description', 'name', 'product')
_STOP_WORDS = ('fits', 'replaces', 'application')


class CatalogSource(Protocol):
    def rows(self) -> list[dict]: ...


def _is_code(token: str) -> bool:
    return bool(_CODE.match(token)) and any(c.isdigit() for c in token)


# --- PDF / plain text -----------------------------------------------------------

def _is_section_header(line: str) -> bool:
    """A section/sub-header line: uppercase-ish words, no part-number, no column
    markers ('#'). Captures 'BEARINGS & BUSHINGS', 'CUMMINS® DIRECT REPLACEMENT'
    — the category/brand context that explains classifier segments."""
    s = line.strip()
    if not s or '#' in s or len(s) > 45:
        return False
    if any(_is_code(t) for t in s.split()):
        return False
    alpha = [c for c in s if c.isalpha()]
    if len(alpha) < 4:
        return False
    return sum(c.isupper() for c in alpha) / len(alpha) >= 0.8


def rows_from_catalog_lines(lines: list[str], *, min_code_len: int = 5
                            ) -> list[dict]:
    """Extract structured rows from catalog text lines.

    Heuristic that survives real layout-mode catalog text: a product line begins
    with a part-number-like code (the catalog's OWN SKU — first column), then an
    optional cross-reference code (OEM #), a word description, and application
    text ('Fits CUMMINS® ...'). All four are captured, plus the running section
    header, so the decoder can correlate segments against fitment/category — not
    just the short description:

      {sku, description, oem, fitment, section}
    """
    out: list[dict] = []
    recent_headers: list[str] = []
    for raw in lines:
        toks = raw.split()
        if _is_section_header(raw):
            h = raw.strip()
            if h not in recent_headers:
                recent_headers.append(h)
                recent_headers[:] = recent_headers[-2:]   # keep last two
            continue
        if len(toks) < 2:
            continue
        sku = toks[0]
        if not (_is_code(sku) and len(sku) >= min_code_len):
            continue
        rest = toks[1:]
        oem = next((t for t in rest if _is_code(t)), '')
        words: list[str] = []
        fit_from = None
        for idx, t in enumerate(rest):
            if t.lower().startswith(_STOP_WORDS) or '®' in t or '™' in t:
                fit_from = idx
                break
            if re.search(r'[A-Za-z]', t) and not _is_code(t):
                words.append(t)
        fitment = ' '.join(rest[fit_from:]) if fit_from is not None else ''
        out.append({'sku': sku, 'description': ' '.join(words), 'oem': oem,
                    'fitment': fitment, 'section': ' | '.join(recent_headers)})
    return out


def extract_pdf_text(path) -> list[str]:
    """Extract text lines from a PDF. Uses pypdf (`pip install '.[pdf]'`).
    For best columnar fidelity on multi-column catalogs, pre-extract with
    `pdftotext -layout` and feed `rows_from_catalog_lines` directly instead."""
    try:
        from pypdf import PdfReader
    except ImportError as e:    # pragma: no cover - environment-dependent
        raise RuntimeError(
            "PDF extraction needs pypdf: pip install '.[pdf]' (or pre-extract "
            "with `pdftotext -layout` and use rows_from_catalog_lines).") from e
    reader = PdfReader(str(path))
    lines: list[str] = []
    for page in reader.pages:
        lines.extend((page.extract_text() or '').splitlines())
    return lines


class PdfCatalogSource:
    def __init__(self, path, *, min_code_len: int = 5) -> None:
        self._path = path
        self._min = min_code_len

    def rows(self) -> list[dict]:
        return rows_from_catalog_lines(extract_pdf_text(self._path),
                                       min_code_len=self._min)


class LineCatalogSource:
    """For already-extracted text (e.g. `pdftotext -layout` output)."""
    def __init__(self, lines: list[str], *, min_code_len: int = 5) -> None:
        self._lines, self._min = lines, min_code_len

    def rows(self) -> list[dict]:
        return rows_from_catalog_lines(self._lines, min_code_len=self._min)


# --- Excel ----------------------------------------------------------------------

def _pick_columns(header: list[str], sample_rows: list[list]) -> tuple[int, int]:
    """Choose (sku_col, desc_col). Prefer header-name match; fall back to
    shape: the most code-like column is the SKU, the longest-text column is
    the description."""
    hl = [str(h or '').strip().lower() for h in header]
    sku_col = next((i for i, h in enumerate(hl)
                    if any(k in h for k in _SKU_HEADERS)), None)
    desc_col = next((i for i, h in enumerate(hl)
                     if any(k in h for k in _DESC_HEADERS)), None)
    ncols = len(header)
    if sku_col is None:
        code_score = [sum(_is_code(str(r[i])) for r in sample_rows if i < len(r))
                      for i in range(ncols)]
        sku_col = max(range(ncols), key=lambda i: code_score[i]) if ncols else 0
    if desc_col is None:
        text_len = [sum(len(str(r[i])) for r in sample_rows
                        if i < len(r) and not _is_code(str(r[i])))
                    for i in range(ncols)]
        desc_col = max((i for i in range(ncols) if i != sku_col),
                       key=lambda i: text_len[i], default=sku_col)
    return sku_col, desc_col


def rows_from_worksheet(matrix: list[list]) -> list[dict]:
    """matrix = list of rows (row 0 = header). Pure; openpyxl-independent."""
    if not matrix:
        return []
    header, body = matrix[0], matrix[1:]
    sku_col, desc_col = _pick_columns(header, body[:50])
    out = []
    for r in body:
        if sku_col >= len(r):
            continue
        sku = str(r[sku_col] or '').strip()
        if not sku:
            continue
        desc = str(r[desc_col]).strip() if desc_col < len(r) and r[desc_col] else ''
        out.append({'sku': sku, 'description': desc})
    return out


class ExcelCatalogSource:
    def __init__(self, path, *, sheet: str | None = None) -> None:
        self._path, self._sheet = path, sheet

    def rows(self) -> list[dict]:
        from openpyxl import load_workbook   # core dependency
        wb = load_workbook(str(self._path), read_only=True, data_only=True)
        ws = wb[self._sheet] if self._sheet else wb.active
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows_from_worksheet(matrix)


# --- Web (HTML tables) ----------------------------------------------------------

class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: list[list[list[str]]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None
        self._table: list[list[str]] | None = None

    def handle_starttag(self, tag, attrs):
        if tag == 'table':
            self._table = []
        elif tag == 'tr' and self._table is not None:
            self._row = []
        elif tag in ('td', 'th') and self._row is not None:
            self._cell = []

    def handle_endtag(self, tag):
        if tag == 'table' and self._table is not None:
            self.tables.append(self._table)
            self._table = None
        elif tag == 'tr' and self._row is not None:
            self._table.append(self._row)
            self._row = None
        elif tag in ('td', 'th') and self._cell is not None:
            self._row.append(' '.join(''.join(self._cell).split()))
            self._cell = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def rows_from_html_tables(html: str) -> list[dict]:
    """Extract {sku, description} from the first usable <table> in an HTML page
    (a vendor catalog/product grid). Stdlib only — no network, no scraper dep."""
    parser = _TableParser()
    parser.feed(html)
    for table in parser.tables:
        if len(table) >= 2 and len(table[0]) >= 2:
            return rows_from_worksheet(table)
    return []


class HtmlCatalogSource:
    def __init__(self, html: str) -> None:
        self._html = html

    def rows(self) -> list[dict]:
        return rows_from_html_tables(self._html)


class HttpHtmlCatalogSource:
    """Fetch a catalog/product-grid page and extract its table rows.

    The fetcher is INJECTED — a `url -> html` callable — so this stays free of
    any network stack (the harness is import-pure). The deployment layer
    provides fetchers: `erp_transport.web_fetch.static_fetcher` (stdlib urllib,
    static pages) or `playwright_fetcher` (JS-rendered grids). Pass whichever
    fits the target site; tests inject a fake fetcher and need no network.
    """
    def __init__(self, url: str, fetcher) -> None:
        if fetcher is None:
            raise ValueError('HttpHtmlCatalogSource needs a fetcher(url)->html; '
                             'see erp_transport.web_fetch')
        self._url, self._fetch = url, fetcher

    def rows(self) -> list[dict]:
        return rows_from_html_tables(self._fetch(self._url))
